from openpyxl import load_workbook

# 读取 excel 文件
workbook = load_workbook('fifa_world_cup.xlsx')
# print(workbook.get_sheet_names())
print(workbook.sheetnames)

# 找到相应的 worksheet
worksheet = workbook['team_info']

# 找到对应国家所在的行
column_name = worksheet['B']
for cell in column_name:
    if cell.value == 'Russia':
        row_index = cell.row
        worksheet['E' + str(row_index)] = 1
    if cell.value == 'Saudi Arabia':
        row_index = cell.row
        cell_target = worksheet.cell(column = 7, row = row_index)
        cell_target.value = 1

# 保存文件
workbook.save('fifa_world_cup.xlsx')